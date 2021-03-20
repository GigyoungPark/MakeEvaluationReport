import win32com.client as win32
from openpyxl import load_workbook

# hwp = win32.Dispatch('HWPFrame.HWPObject') : 보통 이 명령어를 사용하는데 오류가 나서 아래 명령어 사용
hwp = win32.gencache.EnsureDispatch('HWPFrame.HWPObject')
hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")  # 한글파일 열 때 "접근 허용" 없애는 방법, 레지스트지에 등록해서 사용
hwp.XHwpWindows.Item(0).Visible = True  # 작업하는 한글 파일 보이도록 하는 명령어
hwp.Open("D://개인정보보호.hwp", "HWP", "forceopen:true")  # 보고서 작성 양식
hwp.SaveAs("D://개인정보보호_결과.hwp")  # 결과물은 다른 파일에 저장함

# 보고서 양식을 복사해서 기관수 만큼 한글 페이지 생성
# (전체선택 → 복사 → 붙여넣기 → 커서를 페이지 마지막으로 이동 → 페이지 분할 →  붙여넣기 반복)

hwp.Run('SelectAll')
hwp.Run('Copy')
hwp.MovePos(3)
for i in range(1, 159):
    hwp.Run('BreakPage')
    hwp.Run('Paste')

# hwp 안에 필드리스트 조회하는 명령어
# 필드는 한글 파일에 일일이 모두 셋팅해줘야 함 (누름틀 사용)
field_list = [i for i in hwp.GetFieldList().split('\02')]
# print(field_list)
# print(hwp.GetFieldList(1))

# 엑셀 파일 읽기
wb = load_workbook("D://개인정보보호.xlsx", read_only='True', data_only='True')
# 시트 읽기
ws = wb[wb.sheetnames[1]]
print("'" + wb.sheetnames[1] + "' 시트가 활성화 되었습니다.")

# HWP 누름틀에 엑셀 자료 양식에 맞춰서 넣기
for i in range(3, 162):
    row = ws[i]
    print(str(row[3].value) + " 처리 중..")

    hwp.PutFieldText("기관명{{" + str(i-3) + "}}", str(row[3].value))
    hwp.PutFieldText("진단결과{{" + str(i-3) + "}}", format(float(row[4].value), "2.2f"))
    hwp.PutFieldText("공표감점{{" + str(i-3) + "}}", format(float(row[5].value), "1.2f"))
    hwp.PutFieldText("개인정보보호득점{{" + str(i-3) + "}}", round(float(row[6].value)*100, 2))

    print(str(row[3].value) + " 처리 완료")

# HWP 종료
# hwp.HAction.Run("FileClose");
# hwp.HAction.Run("FileQuit");
